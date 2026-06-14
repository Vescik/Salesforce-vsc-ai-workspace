"""XLSX converter (requires ``openpyxl``)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from ai_workspace.knowledge.converters import make_doc


_MAX_ROWS = 200


def convert(path: Path | str) -> dict[str, Any]:
    source = Path(path)
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return make_doc(
            "xlsx",
            source,
            warnings=["openpyxl not installed; .xlsx skipped. Install salesforce-ai-workspace-tools[knowledge]."],
            parse_status="failed",
            degraded=True,
        )

    try:
        workbook = load_workbook(str(source), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return make_doc("xlsx", source, warnings=[f"openpyxl failed to open file: {exc}"], parse_status="failed")

    sections: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    warnings: list[str] = []
    total_rows = 0

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows: list[list[str]] = []
        truncated = False
        for row in sheet.iter_rows(values_only=False):
            if len(rows) >= _MAX_ROWS:
                truncated = True
                break
            cells = [_format_cell(cell) for cell in row]
            if any(c for c in cells):
                rows.append(cells)
        total_rows += len(rows)
        used_range = getattr(sheet, "dimensions", "")
        caption = f"{sheet_name}!{used_range}" if used_range else sheet_name
        sections.append({
            "heading": sheet_name,
            "level": 2,
            "body": "\n".join(" | ".join(r) for r in rows[:50]),
            "anchors": [f"sheet-{sheet_name}"],
        })
        if rows:
            tables.append({"caption": caption, "rows": rows})
        if truncated:
            warnings.append(f"Sheet '{sheet_name}' truncated at {_MAX_ROWS} rows.")

    workbook.close()
    metadata = {
        "library": "openpyxl",
        "sheet_count": len(workbook.sheetnames),
        "row_count": total_rows,
    }
    return make_doc(
        "xlsx",
        source,
        sections=sections,
        tables=tables,
        metadata=metadata,
        warnings=warnings,
        parse_status="partial" if warnings else "ok",
        degraded=False,
    )


def _format_cell(cell: Any) -> str:
    value = cell.value
    if value is None:
        return ""
    fmt = getattr(cell, "number_format", "") or ""
    if hasattr(value, "isoformat") and ("yy" in fmt.lower() or "h" in fmt.lower() or "d" in fmt.lower()):
        return value.isoformat()
    return str(value).strip()
